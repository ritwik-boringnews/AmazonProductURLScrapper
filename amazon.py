from selenium import webdriver  
import time  
import os
import shutil
import subprocess
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook,load_workbook

def colToExcel(col): 
    excelCol = str()
    div = col 
    while div:
        (div, mod) = divmod(div-1, 26) 
        excelCol = chr(mod + 65) + excelCol

    return excelCol



# shutil.copy( "AmazonOutputFileOriginal.xlsx","AmazonOutputFile.xlsx")
INPUT_FILE="amazon_input_file.xlsx"
OUTPUT_FILE="AmazonOutputFile.xlsx"

workbook = load_workbook(filename=INPUT_FILE)
workbookout = load_workbook(filename=OUTPUT_FILE)

sheet = workbook.active
sheetout = workbookout.active

row=1
initial_row=5
driver = webdriver.Chrome()
driver.maximize_window()
j=1

while(sheet["A"+str(row)].value):
    
    try:
        url=sheet["A"+str(row)].value
        driver.get(url)

        try:
            diff_size =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'variation_size_name')))
        except Exception as e:
            print(e)
        # print(diff_size.text)
        # diff_size_list=WebDriverWait(diff_size,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'li')))
        # print(len(diff_size_list))
        TITLE=""
        DESC=""
        MODEL_NUMBER=""
        BRAND=""
        ASIN=""
        CATE=""
        RRP=""
        MRP=""
        IMAGES=[]
        COLORS=""
        PART_NUMBER=""
        DETAILS=[]
        META_DES=""
        META_TIL=""
        PROD_TAG=""
        

        meta_title = driver.find_element_by_xpath("//meta[@name='title']")
        META_TIL=meta_title.get_attribute("content")
        meta_desc = driver.find_element_by_xpath("//meta[@name='description']")
        META_DES = meta_desc.get_attribute("content")
        product_tag =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'nav-progressive-subnav')))
        PROD_TAG=product_tag.text.replace("\n",",")

        get_title = driver.title
        if (len(get_title.split("Amazon.in:"))>1):
            CATE=get_title.split("Amazon.in:")[-1]
        products_title =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'productTitle')))
        TITLE=products_title.text
        try:
            description =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'feature-bullets')))
            ull =WebDriverWait(description,6).until(EC.presence_of_element_located((By.TAG_NAME,'ul')))
            DESC=ull.text
        except:
            DESC=""
        try:
            priceToPay =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.CLASS_NAME,'priceToPay')))
            whole =WebDriverWait(priceToPay,6).until(EC.presence_of_element_located((By.CLASS_NAME,'a-price-whole')))
            RRP=whole.text
        except:
            priceToPay =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.CLASS_NAME,'apexPriceToPay')))
            RRP=priceToPay.text
            

        # decimal =WebDriverWait(priceToPay,6).until(EC.presence_of_element_located((By.CLASS_NAME,'a-price-fraction')))
        try:
            realPrice =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.CLASS_NAME,'basisPrice')))
            mrp =WebDriverWait(realPrice,6).until(EC.presence_of_element_located((By.TAG_NAME,'span')))
            MRP=mrp.text
        except:
            MRP=RRP
      
        try:
            seemore=WebDriverWait(driver,6).until(EC.presence_of_element_located((By.XPATH,'//*[@id="poToggleButton"]/a/span')))
            seemore.click()
        except:
            donothing=1
     
        keep_all_details=[]
        try:
            details_div =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'prodDetails')))
            details =WebDriverWait(details_div,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'tr')))

            for i in details:
                keep_all_details.append(WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'th'))).text.strip())
                keep_all_details.append(WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'td'))).text.strip())
                if ("part number" in i.text or "Part number" in i.text or "Part Number" in i.text):
                    PART_NUMBER=WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'td'))).text
                elif("Brand" in i.text or "BRAND" in i.text or "brand" in i.text):
                    BRAND=WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'td'))).text
                elif("ASIN" in i.text):
                    ASIN=WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'td'))).text
                elif ("model" in i.text or "Model" in i.text ):
                    MODEL_NUMBER=WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'td'))).text
        except:
            details_div =WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'detailBulletsWrapper_feature_div')))
            details =WebDriverWait(details_div,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'li')))

            for i in details:
                keep_all_details.append(i.text.split(':')[0].strip())
                keep_all_details.append(i.text.split(':')[-1].strip())
                if ("part number" in i.text or "Part number" in i.text or "Part Number" in i.text):
                    PART_NUMBER=i.text.split(':')[-1]
                elif ("model" in i.text or "Model" in i.text ):
                    MODEL_NUMBER=i.text.split(':')[-1]
                elif("Brand" in i.text or "BRAND" in i.text or "brand" in i.text):
                    BRAND=i.text.split(':')[-1]
                elif ("ASIN" in i.text):
                    ASIN=i.text.split(':')[-1]
                    
        images_div =WebDriverWait(driver,6).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'item')))
        # images =WebDriverWait(images_div,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'li')))
        image_list=[]
      
        for i in images_div:
            try:
                image=WebDriverWait(i,6).until(EC.presence_of_element_located((By.TAG_NAME,'img')))
                small_image=image.get_attribute('src')
                li=small_image.split('/')
                li1=li[-1].split('.')
                li[-1]=li1[0]+'.'+li1[-1]
                image_list.append("/".join(li))
            except:
                continue
        IMAGES = list(dict.fromkeys(image_list))
        
       
        colors=[]
        try:    
            color_div=WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'tp-inline-twister-dim-values-container')))
            color_list=WebDriverWait(color_div,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'img')))
            for i in color_list:
                colors.append(i.get_attribute('alt'))
            COLORS=",".join(colors)
        except:
            COLORS=",".join(colors)
        # see more
        
        try:
            details=WebDriverWait(driver,6).until(EC.presence_of_element_located((By.ID,'poExpander')))
            tds=WebDriverWait(details,6).until(EC.presence_of_all_elements_located((By.TAG_NAME,'td')))
            for i in tds:
                if(i.text):
                    DETAILS.append(i.text)
        except:
            novalue=1

        if(MRP[0]=='₹'):
            MRP=MRP[1:]
        if(RRP[0]=='₹'):
            RRP=RRP[1:]
        
    

        DETAILS=DETAILS+keep_all_details
        if(len(DETAILS)%2!=0):
            DETAILS.append("")
        
        final_details=[]
        empty_dict = {}
        for i in range(0,len(DETAILS),2):
            empty_dict[DETAILS[i]]=DETAILS[i+1]
            # if("ASIN" in DETAILS[i] or "model" in DETAILS[i] 
            #     or "Model" in DETAILS[i] 
               
            #     or "part number" in DETAILS[i] 
            #     or "Part Number" in DETAILS[i] 
            #     or "Part number" in DETAILS[i] 
            #     or "part Number" in DETAILS[i] 
            #     or "Best Sellers" in DETAILS[i] or "Customer Reviews" in DETAILS[i] or "#" in DETAILS[i]):
                
            #     no=1
            # else:
                
                # final_details.append(DETAILS[i])
                # final_details.append(DETAILS[i+1])
        for key, value in empty_dict.items():
            final_details.append(key)
            final_details.append(value)
        DETAILS=final_details

        filter_details=[]
        for i in range(0,len(DETAILS),2):
            if ("Item model number" in DETAILS[i]):
                MODEL_NUMBER=DETAILS[i+1]
            elif ("Brand" in DETAILS[i]):
                BRAND=DETAILS[i+1]
            elif ("ASIN" in DETAILS[i]):
                ASIN=DETAILS[i+1]
            elif ("Item part number" in DETAILS[i]):
                PART_NUMBER=DETAILS[i+1]

            else:
                filter_details.append(DETAILS[i])
                filter_details.append(DETAILS[i+1])
            
        DETAILS=filter_details
        

        cnt=0
        initial_details_column=58
        while(sheetout[colToExcel(initial_details_column+cnt)+"5"].value):
            for i in range(0,len(DETAILS),2):
                if(sheetout[colToExcel(initial_details_column+cnt)+"5"].value in DETAILS[i]):
                    sheetout[colToExcel(initial_details_column+cnt)+str(row+initial_row)]=DETAILS[i+1]
                    DETAILS[i]="-1"
                    DETAILS[i+1]="-1"
            cnt+=1
            
        for i in range(0,len(DETAILS),2):
            if(DETAILS[i]=='-1' or DETAILS[i+1]=='-1'):
                continue
            sheetout[colToExcel(initial_details_column+cnt)+"5"]=DETAILS[i]
            sheetout[colToExcel(initial_details_column+cnt)+str(row+initial_row)]=DETAILS[i+1]
            cnt+=1

        for i in range(len(IMAGES)):
            if(i>=0 or i<=10):
                sheetout[colToExcel(38+i)+str(row+initial_row)] = IMAGES[i]
       
        MRP = MRP.replace(",", "")

        sheetout["A"+str(row+initial_row)] = j
        sheetout["B"+str(row+initial_row)] = PART_NUMBER
        sheetout["C"+str(row+initial_row)] = TITLE.strip()
        sheetout["D"+str(row+initial_row)] = DESC
        sheetout["F"+str(row+initial_row)] = MODEL_NUMBER
        sheetout["G"+str(row+initial_row)] = BRAND.strip()
        sheetout["H"+str(row+initial_row)] = ASIN.strip()
        sheetout["AB"+str(row+initial_row)] = CATE.strip()
        sheetout["AF"+str(row+initial_row)] = int(float(MRP))
        sheetout["AG"+str(row+initial_row)] = int(float(RRP))
        sheetout["E"+str(row+initial_row)] = PROD_TAG
        sheetout["J"+str(row+initial_row)] = META_TIL
        sheetout["K"+str(row+initial_row)] = META_DES
        sheetout["AC"+str(row+initial_row)] = COLORS
        j=j+1

        print("done : " +url)

    except:
        print('row skipped')
    row+=1
    
    workbookout.save(filename=OUTPUT_FILE)
    
driver.quit()

subprocess.call(('open', OUTPUT_FILE))


